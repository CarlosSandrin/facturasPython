# import re #Regular Expressions
# import os #Trabajar con archivos y carpetas
# import pdfplumber #Trabajar con PDF
# from tkinter import messagebox #interfaz grafica
# from openpyxl import Workbook, load_workbook #Para trabajar con Excel
# from openpyxl.styles import PatternFill, Font
# import os
# import PyPDF2
# from PyPDF2 import PdfReader

# # GET INFO -TELEFONIA FIJA-
# def pdf_a_texto(ruta_completa):
#     texto_completo = ""
#     with open(ruta_completa, "rb") as archivo:
#         lector_pdf = PyPDF2.PdfReader(archivo)
#         for pagina in lector_pdf.pages:
#             texto_completo += pagina.extract_text()
#     return texto_completo

# def getPeriodoFacturadoFija(texto):
#     match = re.search(r"PERIODO FACTURADO\s*(\d{2}/\d{2}/\d{4})\s*al\s*(\d{2}/\d{2}/\d{4})", texto)
#     if match:
#         return f"{match.group(1)} al {match.group(2)}"
#     return None

# def getNroCuentaFija(texto):
#     match = re.search(r"Clave de adhesión al débito automático:\s*(\d+)", texto) 
#     if match:
#         return match.group(1)  
#     return None

# def getFechaVencimientoFija(texto):
#     match = re.search(r"VENCIMIENTO\s*(\d{2}/\d{2}/\d{4})", texto) 
#     if match:
#         return match.group(1)  
#     return None

# def getFechaEmisionFija(texto):
#     match = re.search(r"Fecha deEmisión\s*(\d{2}/\d{2}/\d{4})", texto) 
#     if match:
#         return match.group(1)  
#     return None 

# def getReferenciaPago(texto):
#     match = re.search(r"N°Referencia dePago\s*(\d+)", texto)
#     if match:
#         return match.group(1)  # Retorna el número de referencia encontrado
#     return None

# def getNroClienteFija(texto):
#     """Extrae el número de cliente en formato Cliente Nº: XXXXXXXX."""
#     match = re.search(r"Cliente\s*N[ºo]:?\s*(\d+)", texto)
#     if match:
#         return match.group(1)  
#     return None

# def getNroFacturaFija(texto):
#     match = re.search(r"Factura\s+N°\s+([A-Z]?\d{5}-\d{8})", texto)
#     if match:
#         return match.group(1) 
#     return None

# def getTipoServicioFija(texto):
  
#     palabras_clave = ["Telefonía", "Servicios de Internet"]

#     servicios_encontrados = [servicio.strip('-') for servicio in palabras_clave if servicio in texto]

#     return ", ".join(servicios_encontrados) if servicios_encontrados else "Sin servicio identificado"

# def getMontoTotalDatos(texto):
#     """Busca el monto en todo el texto basado en su formato."""
#     match = re.search(r"\$\s*([\d.]+,\d{2})", texto)
#     if match:
#         return match.group(1) 
#     return None

# def getCargoDelMes(texto):
#     """Busca el monto en todo el texto basado en su formato."""
#     match = re.search(r"CARGOS\s+DEL\s+MES\s+\$\s*([\d.]+,\d{2})", texto)
#     if match:
#         return match.group(1)
#     return None

# def getTipoServicioDatos(texto):
  
#     palabras_clave = ["VPN-IP", "INTERNET SIMETRICO", "INTERNET FULL", "INTERNET SEGURO", "ABONOS"]

#     servicios_encontrados = [servicio.strip('-') for servicio in palabras_clave if servicio in texto]

#     return ", ".join(servicios_encontrados) if servicios_encontrados else "Sin servicio identificado"

# def guardar_datos_excel(datos, archivo_excel):
 
#     if not archivo_excel:
#         print("No se seleccionó un archivo para guardar.")
#         return

#     try:
#         if not os.path.exists(archivo_excel):
#             wb = Workbook()
#             ws = wb.active
#             ws.title = "Facturas"
#             encabezados = ["ORDEN","EMPRESA", "Nº FACTURA", "SERVICIO", "FECHA EMISION","VENCIMIENTO","PERIODO FACTURADO","CARGOS DEL MES", "TOTAL A PAGAR","REFERENCIA PAGO"]
#             ws.append(encabezados)
#             siguiente_orden = 1
#             fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
#             font = Font(bold=True, color="000000")
#             for celda in ws[1]:
#                 celda.fill = fill
#                 celda.font = font
            
#             siguiente_orden = 1
#         else:
#             wb = load_workbook(archivo_excel)
#             ws = wb.active
#             siguiente_orden = ws.max_row 

#         fila_datos = [
#         siguiente_orden,
#         datos.get("empresa"),
#         # datos.get("numero_cliente"),
#         # datos.get("numero_cuenta"),
#         datos.get("numero_factura"),
#         datos.get("servicio"),
#         datos.get("fecha_emision"),
#         datos.get("fecha_vto"),
#         datos.get("periodo_facturado"),
#         datos.get("cargos_mes"),
#         datos.get("monto_total"),
#         datos.get("referencia_pago"),
#         ]
#         ws.append(fila_datos)
#         wb.save(archivo_excel)
#         print(f"Datos guardados en el archivo Excel: {archivo_excel}")

#     except Exception as e:
#         print(f"Error al guardar el archivo: {e}")

# def extraer_info_factura_fija(texto): 
#     """Extrayendo número de factura, fecha y monto total del texto."""
#     datos = {
#         "orden": None,
#         "empresa": "None",
#         # "numero_cliente": None,
#         # "numero_cuenta": None,
#         "numero_factura": None,
#         "servicio": None,
#         "fecha_emision": None,
#         "fecha_vto": None,
#         "periodo_facturado": None,
#         "cargos_mes": None,
#         "monto_total": None,
#         "referencia_pago": None,
#     }
    
#     datos["empresa"]= "TELECOM OTROS"
#     # datos["numero_cliente"] = getNroClienteFija(texto)
#     datos["numero_factura"] = getNroFacturaFija(texto)
#     # datos["numero_cuenta"] =getNroCuentaFija(texto)
#     datos["servicio"] = getTipoServicioFija(texto)
#     datos["cargos_mes"] = getCargoDelMes(texto)
#     datos["periodo_facturado"] = getPeriodoFacturadoFija(texto)
#     datos["fecha_emision"] = getFechaEmisionFija(texto) 
#     datos["fecha_vto"] = getFechaVencimientoFija(texto) 
#     datos["monto_total"] = getMontoTotalDatos(texto)
#     datos["referencia_pago"] = getReferenciaPago(texto)
#     return datos

# def procesar_Telecom_Fija(pdf_path, barra_progreso, archivo_excel): 
    
#     archivos = [archivo for archivo in os.listdir(pdf_path) if archivo.endswith(".pdf")]
#     total_facturas = len(archivos)
#     barra_progreso["maximum"] = total_facturas
#     try:
#         for i, archivo in enumerate(archivos, start=1):
#          ruta_archivo = os.path.join(pdf_path, archivo)

#          texto_extraido = pdf_a_texto(ruta_archivo)
#         #  print(texto_extraido)
#          datos_factura = extraer_info_factura_fija(texto_extraido)
#          guardar_datos_excel(datos_factura, archivo_excel)
        
#          barra_progreso["value"] = i
#          barra_progreso.update_idletasks()
#          print(f"Procesado: {archivo}")
#         messagebox.showinfo("Proceso finalizado", f"Se han procesado: {total_facturas} facturas.")
#     except Exception as e:
#         messagebox.showerror("ERROR", "No se pudo completar el proceso " + {e})

import re #Regular Expressions
import os #Trabajar con archivos y carpetas
import pdfplumber #Trabajar con PDF
from tkinter import messagebox #interfaz grafica
from openpyxl import Workbook, load_workbook #Para trabajar con Excel
from openpyxl.styles import PatternFill, Font
import os
import PyPDF2
from PyPDF2 import PdfReader

# GET INFO -TELEFONIA FIJA-
# def pdf_a_texto(ruta_completa):
#     texto_completo = ""
#     with open(ruta_completa, "rb") as archivo:
#         lector_pdf = PyPDF2.PdfReader(archivo)
#         for pagina in lector_pdf.pages:
#             texto_completo += pagina.extract_text()
#     return texto_completo

def pdf_a_texto(ruta_completa):
    with pdfplumber.open(ruta_completa) as pdf:
     texto = ""
     for pagina in pdf.pages:
        texto += pagina.extract_text()
    return texto

def getPeriodoFacturadoFija(texto):
    match = re.search(r"Período de Facturación\s*(\d{2}/\d{2}/\d{4})\s*al\s*(\d{2}/\d{2}/\d{4})", texto)
    if match:
        return f"{match.group(1)} al {match.group(2)}"
    return None

# def getNroCuentaFija(texto):
#     match = re.search(r"CUENTA\s+N[ºo]:\s*(\d+)", texto) 
#     if match:
#         return match.group(1)  
#     return None

def getFechaVencimientoFija(texto):
    match = re.search(r"Vencimiento:\s*(\d{2}/\d{2}/\d{4})", texto) 
    if match:
        return match.group(1)  
    return None

def getFechaEmisionFija(texto):
    match = re.search(r"Fecha de emisión\s*(\d{2}/\d{2}/\d{4})", texto) 
    if match:
        return match.group(1)  
    return None

def getReferenciaPago(texto):
    match = re.search(r"N°Referencia dePago\s*(\d+)", texto)
    if match:
        return match.group(1)  # Retorna el número de referencia encontrado
    return None 

def getNroClienteFija(texto):
    """Extrae el número de cliente en formato Cliente Nº: XXXXXXXX."""
    match = re.search(r"Número de Cliente\s*(\d{10})", texto)
    if match:
        return match.group(1)  
    return None

def getNroFacturaFija(texto):
    match = re.search(r"Factura\s+N[ºo]:\s*([A-Z]?\d{4}-\d{8})", texto)
    if match:
        return match.group(1) 
    return None

def getTipoServicioFija(texto):
  
    palabras_clave = ["SERVICIO", "INTERNOS", "CANALES", "DDE"]

    servicios_encontrados = [servicio.strip('-') for servicio in palabras_clave if servicio in texto]

    return ", ".join(servicios_encontrados) if servicios_encontrados else "Sin servicio identificado"

def getMontoTotalDatos(texto):
    """Busca el monto en todo el texto basado en su formato."""
    match = re.search(r"\$\s*([\d.]+,\d{2})", texto)
    if match:
        return match.group(1) 
    return None

# def getCargoDelMes(texto):
#     """Busca el monto en todo el texto basado en su formato."""
#     match = re.search(r"CARGOS\s+DEL\s+MES\s+\$\s*([\d.]+,\d{2})", texto)
#     if match:
#         return match.group(1)
#     return None

def getCargoDelMes(texto):
    """Busca el monto en todo el texto basado en su formato."""
    match = re.search(r"CARGOS\s+DEL\s+MES\s+\$\s*([\d.]+,\d{2})", texto)
    if match:
        return match.group(1)
    return None

def getTipoServicioDatos(texto):
  
    palabras_clave = ["VPN-IP", "INTERNET SIMETRICO", "INTERNET FULL", "INTERNET SEGURO", "ABONOS"]

    servicios_encontrados = [servicio.strip('-') for servicio in palabras_clave if servicio in texto]

    return ", ".join(servicios_encontrados) if servicios_encontrados else "Sin servicio identificado"

def guardar_datos_excel(datos, archivo_excel):
 
    if not archivo_excel:
        print("No se seleccionó un archivo para guardar.")
        return

    try:
        if not os.path.exists(archivo_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = "Facturas"
            encabezados = ["ORDEN","EMPRESA","Nº CLIENTE ", "Nº FACTURA", "SERVICIO", "FECHA EMISION","VENCIMIENTO","PERIODO FACTURADO", "TOTAL A PAGAR"]
            ws.append(encabezados)
            siguiente_orden = 1
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
            font = Font(bold=True, color="000000")
            for celda in ws[1]:
                celda.fill = fill
                celda.font = font 
            
            siguiente_orden = 1
        else:
            wb = load_workbook(archivo_excel)
            ws = wb.active
            siguiente_orden = ws.max_row 

        fila_datos = [
        siguiente_orden,
        datos.get("empresa"),
        datos.get("numero_cliente"),
        # datos.get("numero_cuenta"),
        datos.get("numero_factura"),
        datos.get("servicio"),
        datos.get("fecha_emision"),
        datos.get("fecha_vto"),
        datos.get("periodo_facturado"),
        # datos.get("cargos_mes"),
        datos.get("monto_total"),
        datos.get("referencia_pago"),
        ]
        ws.append(fila_datos)
        wb.save(archivo_excel)
        print(f"Datos guardados en el archivo Excel: {archivo_excel}")

    except Exception as e:
        print(f"Error al guardar el archivo: {e}")

def extraer_info_factura_fija(texto): 
    """Extrayendo número de factura, fecha y monto total del texto."""
    datos = {
        "orden": None,
        "empresa": "None",
        "numero_cliente": None,
        # "numero_cuenta": None,
        "numero_factura": None,
        "servicio": None,
        "fecha_emision": None,
        "fecha_vto": None,
        "periodo_facturado": None,
        # "cargos_mes": None,
        "monto_total": None,
        "referencia_pago": None,
    }
    
    datos["empresa"]= "TELECOM OTRO"
    datos["numero_cliente"] = getNroClienteFija(texto)
    datos["numero_factura"] = getNroFacturaFija(texto)
    # datos["numero_cuenta"] =getNroCuentaFija(texto)
    datos["servicio"] = getTipoServicioFija(texto)
    # datos["cargos_mes"] = getCargoDelMes(texto)
    datos["periodo_facturado"] = getPeriodoFacturadoFija(texto)
    datos["fecha_emision"] = getFechaEmisionFija(texto) 
    datos["fecha_vto"] = getFechaVencimientoFija(texto) 
    datos["monto_total"] = getMontoTotalDatos(texto)
    datos["referencia_pago"] = getReferenciaPago(texto)
    return datos

def procesar_Telecom_Otro(pdf_path, barra_progreso, archivo_excel): 
    
    archivos = [archivo for archivo in os.listdir(pdf_path) if archivo.endswith(".pdf")]
    total_facturas = len(archivos)
    barra_progreso["maximum"] = total_facturas
    try:
        for i, archivo in enumerate(archivos, start=1):
         ruta_archivo = os.path.join(pdf_path, archivo)

         texto_extraido = pdf_a_texto(ruta_archivo)
         print(texto_extraido)
         datos_factura = extraer_info_factura_fija(texto_extraido)
         guardar_datos_excel(datos_factura, archivo_excel)
        
         barra_progreso["value"] = i
         barra_progreso.update_idletasks()
         print(f"Procesado: {archivo}")
        messagebox.showinfo("Proceso finalizado", f"Se han procesado: {total_facturas} facturas.")
    except Exception as e:
        messagebox.showerror("ERROR", "No se pudo completar el proceso " + {e})