import re #Regular Expressions
import os #Trabajar con archivos y carpetas
import pdfplumber #Trabajar con PDF
from tkinter import messagebox #interfaz grafica
from openpyxl import Workbook, load_workbook #Para trabajar con Excel
import os
import PyPDF2
from PyPDF2 import PdfReader

# GET INFO -TELEFONIA FIJA-
def pdf_a_texto(ruta_completa):
    with pdfplumber.open(ruta_completa) as pdf:
     texto = ""
     for pagina in pdf.pages:
        texto += pagina.extract_text()
    return texto

def getCargosPeriodoFija(texto):
    match = re.search(r"([\d.,]+)\s*=", texto)
    if match:
        return match.group(1)
    return None

def getNroCuentaFija(texto):
    match = re.search(r"Clave de adhesión al débito automático:\s*(\d+)", texto) 
    if match:
        return match.group(1)  
    return None

def getFechaVencimientoFija(texto):
    match = re.search(r"Vencimiento:\s*(\d{1,2}/\d{1,2}/\d{4})", texto) 
    if match:
        return match.group(1)  
    return None

def getFechaEmisionFija(texto):
    match = re.search(r"Fecha de emisión:\s*(\d{1,2}/\d{1,2}/\d{4})", texto) 
    if match:
        return match.group(1)  
    return None 

def getNroClienteFija(texto):
    """Extrae el número de cliente en formato Cliente Nº: XXXXXXXX."""
    match = re.search(r"Cliente\s*N[ºo]:?\s*(\d+)", texto)
    if match:
        return match.group(1)  
    return None

def getNroFacturaFija(texto):
    match = re.search(r"Factura\s+(\d{4}-\d{8})", texto)
    if match:
        return match.group(1) 
    return None

def getTipoServicioFija(texto):
  
    palabras_clave = ["Telefonía", "Servicios de Internet"]

    servicios_encontrados = [servicio.strip('-') for servicio in palabras_clave if servicio in texto]

    return ", ".join(servicios_encontrados) if servicios_encontrados else "Sin servicio identificado"

def getMontoTotalDatos(texto):
    """Busca el monto en todo el texto basado en su formato."""
    match = re.search(r"\$\s*([\d.]+,\d{2})", texto)
    if match:
        return match.group(1) 
    return None

def getTipoServicioDatos(texto):
  
    palabras_clave = ["VPN-IP", "INTERNET SIMETRICO", "INTERNET FULL", "INTERNET SEGURO"]

    servicios_encontrados = [servicio.strip('-') for servicio in palabras_clave if servicio in texto]

    return ", ".join(servicios_encontrados) if servicios_encontrados else "Sin servicio identificado"

def guardar_datos_excel(datos, archivo_excel):
 
    if not archivo_excel:
        print("No se seleccionó un archivo para guardar.")
        return

    try:
        if not os.path.exists(archivo_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = "Facturas"
            encabezados = ["ORDEN","EMPRESA","CLIENTE", "Nº CUENTA", "Nº FACTURA", "SERVICIO", "FECHA EMISION","VENCIMIENTO","CARGOS PERIODO","TOTAL A PAGAR"]
            ws.append(encabezados)
            siguiente_orden = 1 
        else:
            wb = load_workbook(archivo_excel)
            ws = wb.active
            siguiente_orden = ws.max_row 

        fila_datos = [
        siguiente_orden,
        datos.get("empresa"),
        datos.get("numero_cliente"),
        datos.get("numero_cuenta"),
        datos.get("numero_factura"),
        datos.get("servicio"),
        datos.get("fecha_emision"),
        datos.get("fecha_vto"),
        datos.get("cargos_periodo"),
        datos.get("monto_total"),
        ]
        ws.append(fila_datos)
        wb.save(archivo_excel)
        print(f"Datos guardados en el archivo Excel: {archivo_excel}")

    except Exception as e:
        print(f"Error al guardar el archivo: {e}")

def extraer_info_factura_fija(texto): 
    """Extrayendo número de factura, fecha y monto total del texto."""
    datos = {
        "orden": None,
        "empresa": "None",
        "numero_cliente": None,
        "numero_cuenta": None,
        "numero_factura": None,
        "servicio": None,
        "fecha_emision": None,
        "fecha_vto": None,
        "cargos_periodo": None,
        "monto_total": None,
    }
    
    datos["empresa"]= "TELEFONICA FIJA"
    datos["numero_cliente"] = getNroClienteFija(texto)
    datos["numero_factura"] = getNroFacturaFija(texto)
    datos["numero_cuenta"] =getNroCuentaFija(texto)
    datos["servicio"] = getTipoServicioFija(texto)
    datos["cargos_periodo"] = getCargosPeriodoFija(texto)
    datos["fecha_emision"] = getFechaEmisionFija(texto) 
    datos["fecha_vto"] = getFechaVencimientoFija(texto) 
    datos["monto_total"] = getMontoTotalDatos(texto)
    return datos

def procesar_Telefonica_Fija(pdf_path, barra_progreso, archivo_excel): #DEVUELVE TEXTO PLANO
    
    archivos = [archivo for archivo in os.listdir(pdf_path) if archivo.endswith(".pdf")]
    total_facturas = len(archivos)
    barra_progreso["maximum"] = total_facturas
    try:
        for i, archivo in enumerate(archivos, start=1):
         ruta_archivo = os.path.join(pdf_path, archivo)

         texto_extraido = pdf_a_texto(ruta_archivo)
         datos_factura = extraer_info_factura_fija(texto_extraido)
         guardar_datos_excel(datos_factura, archivo_excel)
        
         barra_progreso["value"] = i
         barra_progreso.update_idletasks()
         print(f"Procesado: {archivo}")
        messagebox.showinfo("Proceso finalizado", f"Se han procesado: {total_facturas} facturas.")
    except:
        messagebox.showerror("ERROR", "No se pudo completar el proceso")





