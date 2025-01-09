import re #Regular Expressions
import os #Trabajar con archivos y carpetas
import pdfplumber #Trabajar con PDF
from tkinter import messagebox
from openpyxl import Workbook, load_workbook #Para trabajar con Excel
import os
import PyPDF2
from PyPDF2 import PdfReader

# GET INFO -TELEFONIA FIJA-

def getCargosPeriodoFija(texto):
    match = re.search(r"([\d.,]+)\s*=", texto)
    if match:
        return match.group(1)
    return None

def getNroCuentaFija(texto):
    match = re.search(r"Clave de adhesión al débito automático:\s*(\d+)", texto) 
    if match:
        return match.group(1)  
    return None

def getFechaVencimientoFija(texto):
    match = re.search(r"Vencimiento:\s*(\d{1,2}/\d{1,2}/\d{4})", texto) 
    if match:
        return match.group(1)  
    return None

def getFechaEmisionFija(texto):
    # Buscar el texto "Período Mensual:" seguido de su contenido
    match = re.search(r"Fecha de emisión:\s*(\d{1,2}/\d{1,2}/\d{4})", texto) #\d{1,2}/\d{1,2}/\d{4}\b"
    if match:
        return match.group(1)  # Devuelve el contenido encontrado
    return None  # Si no se encuentra, devuelve None

def getNroClienteFija(texto):
    """Extrae el número de cliente en formato Cliente Nº: XXXXXXXX."""
    match = re.search(r"Cliente\s*N[ºo]:?\s*(\d+)", texto)
    if match:
        return match.group(1)  
    return None

def getNroLinea(texto):
    match = re.search(r"Línea:\s*(\d+)", texto)
    if match:
        return match.group(1)
    return None

def getNroFacturaFija(texto):
    match = re.search(r"Factura\s+(\d{4}-\d{8})", texto)
    if match:
        return match.group(1) 
    return None

def getTipoServicioFija(texto):
  
    palabras_clave = ["Telefonía", "Servicios de Internet"]

    servicios_encontrados = [servicio.strip('-') for servicio in palabras_clave if servicio in texto]

    return ", ".join(servicios_encontrados) if servicios_encontrados else "Sin servicio identificado"


#////////////////////////////////////////////////////////////////////////////////////////////////////////

# GET INFO -DATOS-
def getNroFacturaDatos(texto):
    lineas = texto.splitlines()  # Divide el texto en líneas
    if len(lineas) >= 3:         # Verifica si hay al menos tres líneas
        return lineas[2]         # Devuelve la tercera línea (índice 2)
    return None  

def getNroClienteDatos(texto):
    """Extrae el número de cliente en formato Cliente: (XXXXXXXX)."""
    match = re.search(r"Cliente:\s*\((\d+)\)", texto)
    if match:
        return match.group(1)  
    return None  

def getFechaByPosicion(texto, posicion):
    """Extrae la fecha que aparece en una posición específica (n-ésima fecha)."""
   
    fechas = re.findall(r"\b\d{1,2}/\d{1,2}/\d{4}\b", texto)
    if len(fechas) >= posicion:  
        return fechas[posicion - 1] 
    return None

def getPeriodoMensualDatos(texto):
   
    match = re.search(r"Período Mensual:\s*(.+)", texto)
    if match:
        return match.group(1)  
    return None  

def getMontoTotalDatos(texto):
    """Busca el monto en todo el texto basado en su formato."""
    match = re.search(r"\$\s*([\d.]+,\d{2})", texto)
    if match:
        return match.group(1) 
    return None

def getTipoServicioDatos(texto):
  
    palabras_clave = ["VPN-IP", "INTERNET SIMETRICO", "INTERNET FULL", "INTERNET SEGURO"]

    servicios_encontrados = [servicio.strip('-') for servicio in palabras_clave if servicio in texto]

    return ", ".join(servicios_encontrados) if servicios_encontrados else "Sin servicio identificado"

def guardar_datos_excel(datos, archivo_excel):
 
    if not archivo_excel:
        print("No se seleccionó un archivo para guardar.")
        return

    try:
        if not os.path.exists(archivo_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = "Facturas"
            encabezados = ["ORDEN","EMPRESA","CLIENTE", "Nº CUENTA", "Nº FACTURA", "SERVICIO", "FECHA EMISION","VENCIMIENTO","CARGOS PERIODO","TOTAL A PAGAR"]
            ws.append(encabezados)
            siguiente_orden = 1 
        else:
            wb = load_workbook(archivo_excel)
            ws = wb.active
            siguiente_orden = ws.max_row 

        fila_datos = [
        siguiente_orden,
        datos.get("empresa"),
        datos.get("numero_cliente"),
        datos.get("numero_cuenta"),
        datos.get("numero_factura"),
        datos.get("servicio"),
        datos.get("fecha_emision"),
        datos.get("fecha_vto"),
        datos.get("cargos_periodo"),
        datos.get("monto_total"),
        ]
        ws.append(fila_datos)
        wb.save(archivo_excel)
        print(f"Datos guardados en el archivo Excel: {archivo_excel}")

    except Exception as e:
        print(f"Error al guardar el archivo: {e}")

#------------------------------------------------------------------------------------


def pdf_a_texto(ruta_completa):
    with pdfplumber.open(ruta_completa) as pdf:
     texto = ""
     for pagina in pdf.pages:
        texto += pagina.extract_text()
    return texto

def extraer_info_factura_datos(texto): 
    """Extrayendo número de factura, fecha y monto total del texto."""
    datos = {
        "orden": None,
        "empresa": "None",
        "numero_cliente": None,
        "numero_cuenta": None,
        "numero_factura": None,
        "servicio": None,
        "fecha_emision": None,
        "fecha_vto": None,
        "cargos_periodo": None,
        "monto_total": None,
    }
    
    datos["empresa"]= "TELEFONICA"
    datos["numero_cliente"] = getNroClienteFija(texto)
    datos["numero_factura"] = getNroFacturaFija(texto)
    datos["numero_cuenta"] =getNroCuentaFija(texto)
    datos["servicio"] = getTipoServicioFija(texto)
    datos["cargos_periodo"] = getCargosPeriodoFija(texto)
    # datos["fecha_emision"] = getFechaByPosicion(texto, 2) PARA DATOS
    datos["fecha_emision"] = getFechaEmisionFija(texto) #PARA TELEFONIA FIJA
    # datos["fecha_vto"] = getFechaByPosicion(texto, 3) PARA FACTURA DE DATOS
    datos["fecha_vto"] = getFechaVencimientoFija(texto) #PARA FACTURA DE TELEFONIA FIJA
    datos["monto_total"] = getMontoTotalDatos(texto)
    datos["numero_linea"] = getNroLinea(texto)
    return datos

def extraer_info_factura_fija(texto): 
    """Extrayendo número de factura, fecha y monto total del texto."""
    datos = {
        "orden": None,
        "empresa": "None",
        "numero_cliente": None,
        "numero_cuenta": None,
        "numero_factura": None,
        "servicio": None,
        "fecha_emision": None,
        "fecha_vto": None,
        "cargos_periodo": None,
        "monto_total": None,
    }
    
    datos["empresa"]= "TELEFONICA"
    datos["numero_cliente"] = getNroClienteFija(texto)
    datos["numero_factura"] = getNroFacturaFija(texto)
    datos["numero_cuenta"] =getNroCuentaFija(texto)
    datos["servicio"] = getTipoServicioFija(texto)
    datos["cargos_periodo"] = getCargosPeriodoFija(texto)
    # datos["fecha_emision"] = getFechaByPosicion(texto, 2) PARA DATOS
    datos["fecha_emision"] = getFechaEmisionFija(texto) #PARA TELEFONIA FIJA
    # datos["fecha_vto"] = getFechaByPosicion(texto, 3) PARA FACTURA DE DATOS
    datos["fecha_vto"] = getFechaVencimientoFija(texto) #PARA FACTURA DE TELEFONIA FIJA
    datos["monto_total"] = getMontoTotalDatos(texto)
    datos["numero_linea"] = getNroLinea(texto)
    return datos

def renameDoc(ruta_origen, numero_cliente, periodo):
    """
    Renombra un archivo PDF con el formato especificado y lo mueve a la carpeta 'outputs'.

    Args:
        ruta_origen (str): Ruta original del archivo PDF.
        numero_cliente (str): Número de cliente extraído.
        periodo (str): Período mensual extraído.
    """
    # Ruta de destino
    ruta_destino = "./outputs/Datos/"
    
    # Crear la carpeta de destino si no existe
    if not os.path.exists(ruta_destino):
        os.makedirs(ruta_destino)  # Crea la carpeta

    # Construir el nuevo nombre del archivo
    nuevo_nombre = f"{numero_cliente}_{periodo}_DATOS.pdf"
    nuevo_nombre = (
        nuevo_nombre.replace(" ", "_")  # Reemplazar espacios por guiones bajos
        .replace("/", "-")              # Reemplazar barras por guiones
    )
    ruta_completa_destino = os.path.join(ruta_destino, nuevo_nombre)

    # Renombrar el archivo
    print(f"Renombrando archivo:\n - Desde: {ruta_origen}\n - Hacia: {ruta_completa_destino}")
    if os.path.exists(ruta_origen):
       os.rename(ruta_origen, ruta_destino)
       print(f"Archivo renombrado y movido a: {ruta_completa_destino}")
    else:
        print(f"Error: No se encontró el archivo de origen: {ruta_origen}")

def procesar_Telecom_Movil(pdf_path, barra_progreso, archivo_excel): #DEVUELVE TEXTO PLANO
    
    archivos = [archivo for archivo in os.listdir(pdf_path) if archivo.endswith(".pdf")]
    total_facturas = len(archivos)
    barra_progreso["maximum"] = total_facturas
    try:
        for i, archivo in enumerate(archivos, start=1):
         ruta_archivo = os.path.join(pdf_path, archivo)

         texto_extraido = pdf_a_texto(ruta_archivo)
         datos_factura = extraer_info_factura_fija(texto_extraido)
         guardar_datos_excel(datos_factura, archivo_excel)
        
         barra_progreso["value"] = i
         barra_progreso.update_idletasks()
         print(f"Procesado: {archivo}")
        messagebox.showinfo("Proceso finalizado", f"Se han procesado: {total_facturas} facturas.")
    except:
        messagebox.showerror("ERROR", "No se pudo completar el proceso")

    
    












